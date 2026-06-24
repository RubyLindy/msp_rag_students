import argparse
import os
from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

from app.rag.pipeline import answer


def run_batch(input_file, output_file=None):
    if not input_file.endswith(".txt"):
        raise ValueError(f"Input file must be a .txt file, got: '{input_file}'")

    with open(input_file, "r", encoding="utf-8") as f:
        questions = [line.strip() for line in f if line.strip()]

    print(f"Running {len(questions)} questions from '{input_file}'...\n")

    rows = []
    for i, question in enumerate(questions, 1):
        print(f"[{i}/{len(questions)}] {question}")
        try:
            result = answer(question)
            sources_text = "\n".join(
                f"{s['file']} (distance: {s['distance']})" for s in result["sources"]
            )
            rows.append({
                "question": question,
                "answer": result["answer"],
                "sources": sources_text,
                "error": ""
            })
            print(f"  -> Done ({len(result['sources'])} sources)")
            for s in result["sources"]:
                print(f"     {s['file']} (distance: {s['distance']})")
            print()
        except Exception as e:
            rows.append({
                "question": question,
                "answer": "",
                "sources": "",
                "error": str(e)
            })
            print(f"  -> Error: {e}\n")

    # Build output path
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"logs/batch_{base}_{timestamp}.xlsx"

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["#", "Question", "Answer", "Sources", "Error"]
    col_widths = [5, 40, 60, 40, 30]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=i - 1).alignment = cell_alignment
        ws.cell(row=i, column=2, value=row["question"]).alignment = cell_alignment
        ws.cell(row=i, column=3, value=row["answer"]).alignment = cell_alignment
        ws.cell(row=i, column=4, value=row["sources"]).alignment = cell_alignment
        ws.cell(row=i, column=5, value=row["error"]).alignment = cell_alignment

        # Alternate row shading
        if i % 2 == 0:
            fill = PatternFill(fill_type="solid", fgColor="EFF6FF")
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = fill

    wb.save(output_file)
    print(f"Results saved to '{output_file}'")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run MSP RAG on a list of questions without the UI")
    parser.add_argument("input", help="Path to a .txt file with one question per line")
    parser.add_argument("--output", help="Path to output .xlsx file (optional, auto-generated if omitted)")
    args = parser.parse_args()

    run_batch(args.input, args.output)

## Make the JSON output a excel file for the students
## Gemma4 and openai